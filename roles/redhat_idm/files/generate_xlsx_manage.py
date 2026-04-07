#!/usr/bin/env python3
"""
IPA Manage Report Generator — generates a formatted XLSX from a manage operation result.

Usage:
    python3 generate_xlsx_manage.py <input.json> <output.xlsx>

Input JSON format:
    {
        "operation": "creation",
        "meta": {"client_id": "1141", "ipa_domain": "...", "ipa_server": "...", "generated_at": "..."},
        "rows": [{"username": "jdoe", "status": "Created", "password": "...", ...}]
    }
"""

import json
import sys
import argparse

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ── Palette (identical to generate_xlsx_report.py) ───────────────────────────

FILL_HEADER   = PatternFill("solid", fgColor="1F4E79")   # Dark blue
FILL_SECTION  = PatternFill("solid", fgColor="2E75B6")   # Medium blue
FILL_OK       = PatternFill("solid", fgColor="E8F5E9")   # Light green
FILL_WARNING  = PatternFill("solid", fgColor="FFF9C4")   # Yellow
FILL_ERROR    = PatternFill("solid", fgColor="FFD7D7")   # Red
FILL_WHITE    = PatternFill("solid", fgColor="FFFFFF")

FONT_HEADER   = Font(bold=True, color="FFFFFF", size=11)
FONT_SECTION  = Font(bold=True, color="FFFFFF", size=11)
FONT_BOLD     = Font(bold=True)
FONT_NORMAL   = Font()

TABLE_STYLE   = "TableStyleMedium9"

THIN_BORDER = Border(bottom=Side(style="thin", color="CCCCCC"))

# ── Operation schemas ─────────────────────────────────────────────────────────

SCHEMAS = {
    "creation": {
        "title":   "User Creation",
        "headers": ["Username", "Firstname", "Lastname", "Email", "Groups", "Status", "Password", "Client ID", "IPA Domain"],
        "keys":    ["username", "firstname", "lastname", "email", "groups", "status", "password", "client_id", "ipa_domain"],
    },
    "password_reset": {
        "title":   "Password Reset",
        "headers": ["Username", "Status", "Password", "Client ID", "IPA Domain"],
        "keys":    ["username", "status", "password", "client_id", "ipa_domain"],
    },
    "enabling": {
        "title":   "User Enabling",
        "headers": ["Username", "Status", "Password", "Client ID", "IPA Domain"],
        "keys":    ["username", "status", "password", "client_id", "ipa_domain"],
    },
    "deletion": {
        "title":   "User Deletion",
        "headers": ["Username", "Status", "Client ID", "IPA Domain"],
        "keys":    ["username", "status", "client_id", "ipa_domain"],
    },
    "disabling": {
        "title":   "User Disabling",
        "headers": ["Username", "Status", "Client ID", "IPA Domain"],
        "keys":    ["username", "status", "client_id", "ipa_domain"],
    },
    "add_group": {
        "title":   "Add User to Group",
        "headers": ["Username", "Group", "Status", "Client ID", "IPA Domain"],
        "keys":    ["username", "group", "status", "client_id", "ipa_domain"],
    },
    "remove_group": {
        "title":   "Remove User from Group",
        "headers": ["Username", "Group", "Status", "Client ID", "IPA Domain"],
        "keys":    ["username", "group", "status", "client_id", "ipa_domain"],
    },
    "add_pubkey": {
        "title":   "SSH Public Key Injection",
        "headers": ["Username", "Status", "IPA Server"],
        "keys":    ["username", "status", "ipa_server"],
    },
}

# Status values → fill color
STATUS_OK      = {"Created", "Reset", "Enabled", "Deleted", "Disabled", "Added", "Removed", "Injected"}
STATUS_WARNING = {"Already exists", "Already enabled", "Already disabled",
                  "Already in group", "Already removed", "No key", "Not found"}


def get_status_fill(status):
    s = str(status or "")
    if s in STATUS_OK:
        return FILL_OK
    if s in STATUS_WARNING:
        return FILL_WARNING
    if s.startswith("Error"):
        return FILL_ERROR
    return None


# ── Styling helpers ───────────────────────────────────────────────────────────

def write_meta_section(ws, meta, operation_title):
    """Write a styled meta block (rows 1-6) and return the next row index (7)."""
    # Row 1 — operation title bar
    ws.row_dimensions[1].height = 22
    cell = ws.cell(row=1, column=1, value=operation_title)
    cell.fill = FILL_SECTION
    cell.font = FONT_SECTION
    cell.alignment = Alignment(horizontal="left", vertical="center")
    # Merge across all columns (estimate wide enough)
    ws.merge_cells("A1:I1")

    def kv(row, key, value):
        ka = ws.cell(row=row, column=1, value=key)
        va = ws.cell(row=row, column=2, value=value)
        ka.font = FONT_BOLD
        ka.border = THIN_BORDER
        va.border = THIN_BORDER
        ws.column_dimensions["A"].width = 18
        ws.column_dimensions["B"].width = 40

    kv(2, "Generated at", meta.get("generated_at", "-"))
    kv(3, "IPA Domain",   meta.get("ipa_domain", "-"))
    kv(4, "IPA Server",   meta.get("ipa_server", "-"))
    kv(5, "Client ID",    meta.get("client_id", "-"))
    # Row 6 — spacer
    return 7


def write_header_row(ws, headers, row):
    """Write a styled header row at the given row index."""
    ws.row_dimensions[row].height = 20
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)


def apply_table(ws, table_name, headers, header_row):
    """Add an Excel Table from header_row to last row."""
    last_row = ws.max_row
    if last_row <= header_row:
        return   # No data rows — skip table (table with 0 data rows crashes openpyxl)
    ref = f"A{header_row}:{get_column_letter(len(headers))}{last_row}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=TABLE_STYLE,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)
    ws.freeze_panes = f"A{header_row + 1}"


def auto_width(ws, min_w=8, max_w=55):
    """Auto-size columns based on content, skipping merged cells."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=0,
        )
        current = ws.column_dimensions[col_letter].width
        new_w = min(max_w, max(min_w, max_len + 2))
        # Don't shrink meta columns A/B which were set explicitly
        ws.column_dimensions[col_letter].width = max(current, new_w)


# ── Main sheet builder ────────────────────────────────────────────────────────

def build_sheet(wb, operation, meta, rows):
    schema = SCHEMAS.get(operation)
    if schema is None:
        ws = wb.active or wb.create_sheet("Result")
        ws.cell(row=1, column=1, value=f"Unknown operation: {operation}")
        return

    headers  = schema["headers"]
    keys     = schema["keys"]
    title    = schema["title"]
    sh_title = title[:31]   # Excel sheet name limit

    ws = wb.active
    ws.title = sh_title

    # Meta section (rows 1-6) — data starts at row 7
    data_start = write_meta_section(ws, meta, title)
    header_row = data_start          # row 7
    first_data = data_start + 1      # row 8

    write_header_row(ws, headers, header_row)

    for i, row in enumerate(rows, start=first_data):
        values = [str(row.get(k, "-") or "-") for k in keys]
        status = row.get("status", "")
        fill   = get_status_fill(status)

        for col, val in enumerate(values, start=1):
            cell = ws.cell(row=i, column=col, value=val)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

    # Safe table name: alphanumeric only
    table_name = "".join(c for c in operation if c.isalnum()).capitalize()
    apply_table(ws, table_name, headers, header_row)
    auto_width(ws)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate XLSX manage report from JSON.")
    parser.add_argument("input_json",  help="Path to the operation result JSON file")
    parser.add_argument("output_xlsx", help="Path to write the output XLSX file")
    args = parser.parse_args()

    try:
        with open(args.input_json, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError) as exc:
        print(f"ERROR: Cannot read input JSON: {exc}", file=sys.stderr)
        sys.exit(1)

    operation = data.get("operation", "unknown")
    meta      = data.get("meta", {})
    rows      = data.get("rows", [])

    wb = Workbook()
    build_sheet(wb, operation, meta, rows)

    try:
        wb.save(args.output_xlsx)
        username = rows[0].get("username", "?") if rows else "?"
        print(f"Report saved: {args.output_xlsx} (operation={operation}, user={username})")
    except OSError as exc:
        print(f"ERROR: Cannot write output XLSX: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
