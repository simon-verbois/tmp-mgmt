#!/usr/bin/env python3
"""
IPA Report Generator — generates a formatted XLSX from FreeIPA JSON data.

Usage:
    python3 generate_xlsx_report.py <input.json> <output.xlsx>
"""

import json
import sys
import argparse
from datetime import datetime, timezone, timedelta

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# ── Palette ──────────────────────────────────────────────────────────────────

FILL_HEADER       = PatternFill("solid", fgColor="1F4E79")   # Dark blue
FILL_SECTION      = PatternFill("solid", fgColor="2E75B6")   # Medium blue
FILL_DISABLED     = PatternFill("solid", fgColor="FFD7D7")   # Red (disabled users)
FILL_EXPIRED      = PatternFill("solid", fgColor="FFE0B2")   # Orange (expired passwords)
FILL_EXPIRING     = PatternFill("solid", fgColor="FFF9C4")   # Yellow (expiring soon)
FILL_ALERT        = PatternFill("solid", fgColor="FCE4D6")   # Light red (alert rows)
FILL_OK           = PatternFill("solid", fgColor="E8F5E9")   # Light green (OK rows)
FILL_WHITE        = PatternFill("solid", fgColor="FFFFFF")

FONT_HEADER       = Font(bold=True, color="FFFFFF", size=11)
FONT_SECTION      = Font(bold=True, color="FFFFFF", size=11)
FONT_BOLD         = Font(bold=True)
FONT_ALERT        = Font(bold=True, color="9C0006")
FONT_NORMAL       = Font()

TABLE_STYLE       = "TableStyleMedium9"

THIN_BORDER = Border(
    bottom=Side(style="thin", color="CCCCCC"),
)

# ── Helpers ───────────────────────────────────────────────────────────────────

def get_val(obj, key, default="-"):
    """Return first element of an IPA attribute list, or default."""
    val = obj.get(key)
    if val is None:
        return default
    if isinstance(val, list):
        if not val:
            return default
        v = val[0]
        if isinstance(v, dict) and "__datetime__" in v:
            return _fmt_date(v["__datetime__"])
        return str(v) if v not in (None, "") else default
    return str(val) if val not in (None, "") else default


def get_list(obj, key, sep=" | ", default="-"):
    """Join a multi-value IPA attribute list."""
    val = obj.get(key)
    if not val:
        return default
    parts = []
    for v in val:
        if isinstance(v, dict) and "__datetime__" in v:
            parts.append(_fmt_date(v["__datetime__"]))
        elif v not in (None, ""):
            parts.append(str(v))
    return sep.join(parts) if parts else default


def get_date(obj, key, default="-"):
    """Return a formatted date string from an IPA datetime attribute."""
    val = obj.get(key)
    if not val or not isinstance(val, list) or not val[0]:
        return default
    v = val[0]
    raw = v["__datetime__"] if isinstance(v, dict) and "__datetime__" in v else str(v)
    return _fmt_date(raw)


def _fmt_date(raw):
    """Convert IPA date YYYYMMDDHHMMSSZ → DD/MM/YYYY HH:MM:SS."""
    try:
        dt = datetime.strptime(str(raw).strip(), "%Y%m%d%H%M%SZ")
        return dt.strftime("%d/%m/%Y %H:%M:%S")
    except (ValueError, TypeError):
        return str(raw)


def _parse_dt(obj, key):
    """Parse IPA datetime attribute to a timezone-aware datetime, or None."""
    try:
        val = obj.get(key)
        if not val or not isinstance(val, list) or not val[0]:
            return None
        v = val[0]
        raw = v["__datetime__"] if isinstance(v, dict) and "__datetime__" in v else str(v)
        return datetime.strptime(str(raw).strip(), "%Y%m%d%H%M%SZ").replace(tzinfo=timezone.utc)
    except (ValueError, TypeError, KeyError):
        return None


def is_expired(obj, key):
    dt = _parse_dt(obj, key)
    return dt is not None and dt < datetime.now(timezone.utc)


def is_expiring_soon(obj, key, days=30):
    dt = _parse_dt(obj, key)
    if dt is None:
        return False
    now = datetime.now(timezone.utc)
    return now <= dt <= now + timedelta(days=days)


def is_disabled(obj):
    val = obj.get("nsaccountlock", False)
    if isinstance(val, list):
        val = val[0] if val else False
    if isinstance(val, str):
        return val.lower() in ("true", "1")
    return bool(val)


def get_results(data, key):
    """Safely extract the result list from IPA JSON."""
    try:
        r = data[key]["result"]["result"]
        return r if isinstance(r, list) else None
    except (KeyError, TypeError):
        return None


def extract_service_host(principal):
    """Extract host from a Kerberos service principal like 'HTTP/host.domain.com@REALM'."""
    try:
        return principal.split("/", 1)[1].split("@")[0]
    except (IndexError, AttributeError):
        return "-"


# ── Styling helpers ───────────────────────────────────────────────────────────

def write_header(ws, headers):
    """Write a styled header row (row 1)."""
    ws.row_dimensions[1].height = 20
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=False)


def apply_table(ws, table_name, n_cols):
    """Add an Excel Table with auto-filter and freeze header row."""
    if ws.max_row < 2:
        return
    ref = f"A1:{get_column_letter(n_cols)}{ws.max_row}"
    tbl = Table(displayName=table_name, ref=ref)
    tbl.tableStyleInfo = TableStyleInfo(
        name=TABLE_STYLE,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tbl)
    ws.freeze_panes = "A2"


def auto_width(ws, min_w=8, max_w=55):
    """Auto-size columns based on content."""
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[col_letter].width = min(max_w, max(min_w, max_len + 2))


def fill_row(ws, row_idx, values, fill=None):
    """Write a data row and optionally apply a fill to all its cells."""
    for col, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col, value=val)
        if fill:
            cell.fill = fill
        cell.alignment = Alignment(horizontal="left", vertical="center")


def error_sheet(wb, title, message):
    """Create a minimal error sheet when data is unavailable."""
    ws = wb.create_sheet(title=title)
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 65
    write_header(ws, ["Status", "Detail"])
    ws.cell(row=2, column=1, value="ERROR").fill = FILL_ALERT
    ws.cell(row=2, column=1).font = FONT_ALERT
    ws.cell(row=2, column=2, value=message)
    return ws


# ── Sheet builders ────────────────────────────────────────────────────────────

def build_users(wb, data):
    """Sheet: Users — returns (disabled_count, expired_count, expiring_count)."""
    rows = get_results(data, "users")
    if rows is None:
        error_sheet(wb, "Users", "User data could not be fetched from IPA.")
        return 0, 0, 0

    ws = wb.create_sheet(title="Users")
    headers = [
        "Username", "Firstname", "Lastname", "Email", "Disabled",
        "Groups", "Home Folder", "Shell", "Password Expiration",
        "Indirect Groups", "Indirect Sudo Rules", "Indirect HBAC Rules",
    ]
    write_header(ws, headers)

    disabled_count = expired_count = expiring_count = 0

    for i, u in enumerate(rows, start=2):
        disabled = is_disabled(u)
        expired = is_expired(u, "krbpasswordexpiration")
        expiring = is_expiring_soon(u, "krbpasswordexpiration", days=30)

        values = [
            get_val(u, "uid"),
            get_val(u, "givenname"),
            get_val(u, "sn"),
            get_val(u, "mail"),
            "Yes" if disabled else "No",
            get_list(u, "memberof_group"),
            get_val(u, "homedirectory"),
            get_val(u, "loginshell"),
            get_date(u, "krbpasswordexpiration"),
            get_list(u, "memberofindirect_group"),
            get_list(u, "memberofindirect_sudorule"),
            get_list(u, "memberofindirect_hbacrule"),
        ]

        if disabled:
            fill_row(ws, i, values, FILL_DISABLED)
            disabled_count += 1
        elif expired:
            fill_row(ws, i, values, FILL_EXPIRED)
            expired_count += 1
        elif expiring:
            fill_row(ws, i, values, FILL_EXPIRING)
            expiring_count += 1
        else:
            fill_row(ws, i, values)

    apply_table(ws, "Users", len(headers))
    auto_width(ws)
    return disabled_count, expired_count, expiring_count


def build_groups(wb, data):
    rows = get_results(data, "groups")
    if rows is None:
        error_sheet(wb, "User Groups", "User group data could not be fetched from IPA.")
        return

    ws = wb.create_sheet(title="User Groups")
    headers = [
        "Group Name", "Description", "Member Users", "Member Groups",
        "Member Of Groups", "Member Of HBAC Rules",
        "Indirect Sudo Rules", "Indirect HBAC Rules",
    ]
    write_header(ws, headers)

    for i, g in enumerate(rows, start=2):
        fill_row(ws, i, [
            get_val(g, "cn"),
            get_val(g, "description"),
            get_list(g, "member_user"),
            get_list(g, "member_group"),
            get_list(g, "memberof_group"),
            get_list(g, "memberof_hbacrule"),
            get_list(g, "memberofindirect_sudorule"),
            get_list(g, "memberofindirect_hbacrule"),
        ])

    apply_table(ws, "UserGroups", len(headers))
    auto_width(ws)


def build_hosts(wb, data):
    rows = get_results(data, "hosts")
    if rows is None:
        error_sheet(wb, "Hosts", "Host data could not be fetched from IPA.")
        return

    ws = wb.create_sheet(title="Hosts")
    headers = [
        "Hostname", "Operating System", "Host Groups",
        "Sudo Rules", "HBAC Rules",
        "Indirect Sudo Rules", "Indirect HBAC Rules",
    ]
    write_header(ws, headers)

    for i, h in enumerate(rows, start=2):
        fill_row(ws, i, [
            get_val(h, "fqdn"),
            get_val(h, "os"),
            get_list(h, "memberof_hostgroup"),
            get_list(h, "memberof_sudorule"),
            get_list(h, "memberof_hbacrule"),
            get_list(h, "memberofindirect_sudorule"),
            get_list(h, "memberofindirect_hbacrule"),
        ])

    apply_table(ws, "Hosts", len(headers))
    auto_width(ws)


def build_host_groups(wb, data):
    rows = get_results(data, "host_groups")
    if rows is None:
        error_sheet(wb, "Host Groups", "Host group data could not be fetched from IPA.")
        return

    ws = wb.create_sheet(title="Host Groups")
    headers = [
        "Group Name", "Description", "Member Hosts", "Member Groups",
        "Member Of Sudo Rules", "Member Of HBAC Rules",
    ]
    write_header(ws, headers)

    for i, g in enumerate(rows, start=2):
        fill_row(ws, i, [
            get_val(g, "cn"),
            get_val(g, "description"),
            get_list(g, "member_host"),
            get_list(g, "member_hostgroup"),
            get_list(g, "memberof_sudorule"),
            get_list(g, "memberof_hbacrule"),
        ])

    apply_table(ws, "HostGroups", len(headers))
    auto_width(ws)


def build_sudo_rules(wb, data):
    rows = get_results(data, "sudo_rules")
    if rows is None:
        error_sheet(wb, "Sudo Rules", "Sudo rule data could not be fetched from IPA.")
        return

    ws = wb.create_sheet(title="Sudo Rules")
    headers = [
        "Rule Name", "Description", "Enabled", "Options",
        "Allowed Commands", "Allowed Command Groups", "Run As External",
        "Users", "User Groups", "Hosts", "Host Groups",
    ]
    write_header(ws, headers)

    for i, r in enumerate(rows, start=2):
        enabled = get_val(r, "ipaenabledflag")
        fill = FILL_DISABLED if enabled.lower() in ("false", "no") else None
        fill_row(ws, i, [
            get_val(r, "cn"),
            get_val(r, "description"),
            enabled,
            get_list(r, "ipasudoopt"),
            get_list(r, "ipasudocmd"),
            get_list(r, "ipasudocmdgrp"),
            get_val(r, "ipasudorunas_external"),
            get_list(r, "memberuser_user"),
            get_list(r, "memberuser_group"),
            get_list(r, "memberhost_host"),
            get_list(r, "memberhost_hostgroup"),
        ], fill=fill)

    apply_table(ws, "SudoRules", len(headers))
    auto_width(ws)


def build_hbac_rules(wb, data):
    rows = get_results(data, "hbac_rules")
    if rows is None:
        error_sheet(wb, "HBAC Rules", "HBAC rule data could not be fetched from IPA.")
        return

    ws = wb.create_sheet(title="HBAC Rules")
    headers = [
        "Rule Name", "Description", "Enabled",
        "HBAC Services", "HBAC Service Groups",
        "Users", "User Groups", "Hosts", "Host Groups",
    ]
    write_header(ws, headers)

    for i, r in enumerate(rows, start=2):
        enabled = get_val(r, "ipaenabledflag")
        fill = FILL_DISABLED if enabled.lower() in ("false", "no") else None
        fill_row(ws, i, [
            get_val(r, "cn"),
            get_val(r, "description"),
            enabled,
            get_list(r, "memberservice_hbacservice"),
            get_list(r, "memberservice_hbacservicegroup"),
            get_list(r, "memberuser_user"),
            get_list(r, "memberuser_group"),
            get_list(r, "memberhost_host"),
            get_list(r, "memberhost_hostgroup"),
        ], fill=fill)

    apply_table(ws, "HBACRules", len(headers))
    auto_width(ws)


def build_pw_policies(wb, data):
    rows = get_results(data, "pw_policies")
    if rows is None:
        error_sheet(wb, "Password Policies", "Password policy data could not be fetched from IPA.")
        return

    ws = wb.create_sheet(title="Password Policies")
    headers = [
        "Policy Group", "Max Lifetime (days)", "Min Lifetime (hours)",
        "History Size", "Character Classes", "Min Length", "Priority",
        "Max Failures", "Failure Reset Interval (s)", "Lockout Duration (s)",
        "Grace Login Limit",
    ]
    write_header(ws, headers)

    for i, p in enumerate(rows, start=2):
        fill_row(ws, i, [
            get_val(p, "cn"),
            get_val(p, "krbmaxpwdlife"),
            get_val(p, "krbminpwdlife"),
            get_val(p, "krbpwdhistorylength"),
            get_val(p, "krbpwdmincategories"),
            get_val(p, "krbpwdminlength"),
            get_val(p, "krbpwdpolicypriority"),
            get_val(p, "krbpwdmaxfailure"),
            get_val(p, "krbpwdfailurecountinterval"),
            get_val(p, "krbpwdlockoutduration"),
            get_val(p, "krbpwdgracelimit"),
        ])

    apply_table(ws, "PwPolicies", len(headers))
    auto_width(ws)


def build_roles(wb, data):
    rows = get_results(data, "roles")
    if rows is None:
        error_sheet(wb, "Roles", "Role data could not be fetched from IPA.")
        return

    ws = wb.create_sheet(title="Roles")
    headers = [
        "Role Name", "Description",
        "Member Users", "Member Groups", "Member Services",
        "Privileges",
    ]
    write_header(ws, headers)

    for i, r in enumerate(rows, start=2):
        fill_row(ws, i, [
            get_val(r, "cn"),
            get_val(r, "description"),
            get_list(r, "member_user"),
            get_list(r, "member_group"),
            get_list(r, "member_service"),
            get_list(r, "memberof_privilege"),
        ])

    apply_table(ws, "Roles", len(headers))
    auto_width(ws)


def build_services(wb, data):
    rows = get_results(data, "services")
    if rows is None:
        error_sheet(wb, "Services", "Service data could not be fetched from IPA.")
        return

    ws = wb.create_sheet(title="Services")
    headers = [
        "Service Principal", "Host", "Has Certificate",
        "Managed By", "Kerberos Aliases",
    ]
    write_header(ws, headers)

    for i, s in enumerate(rows, start=2):
        principal = get_val(s, "krbprincipalname")
        host = extract_service_host(principal) if principal != "-" else "-"
        has_cert = "Yes" if s.get("usercertificate") else "No"
        fill_row(ws, i, [
            principal,
            host,
            has_cert,
            get_list(s, "managedby_host"),
            get_list(s, "krbprincipalname"),
        ])

    apply_table(ws, "Services", len(headers))
    auto_width(ws)


def build_dns_zones(wb, data):
    rows = get_results(data, "dns_zones")
    if rows is None:
        error_sheet(wb, "DNS Zones", "DNS zone data could not be fetched from IPA.")
        return

    ws = wb.create_sheet(title="DNS Zones")
    headers = [
        "Zone Name", "Authoritative NS", "Admin Email",
        "Serial", "Refresh", "Expire", "Min TTL",
        "Active", "Allow Queries", "Allow Transfer",
    ]
    write_header(ws, headers)

    for i, z in enumerate(rows, start=2):
        active = get_val(z, "idnszoneactive")
        fill = FILL_DISABLED if active.lower() in ("false", "no") else None
        fill_row(ws, i, [
            get_val(z, "idnsname"),
            get_val(z, "idnssoamname"),
            get_val(z, "idnssoarname"),
            get_val(z, "idnssoaserial"),
            get_val(z, "idnssoarefresh"),
            get_val(z, "idnssoaexpire"),
            get_val(z, "idnssoaretry"),
            active,
            get_val(z, "idnsallowquery"),
            get_val(z, "idnsallowtransfer"),
        ], fill=fill)

    apply_table(ws, "DNSZones", len(headers))
    auto_width(ws)


def build_hbac_services(wb, data):
    rows = get_results(data, "hbac_services")
    if rows is None:
        error_sheet(wb, "HBAC Services", "HBAC service data could not be fetched from IPA.")
        return

    ws = wb.create_sheet(title="HBAC Services")
    headers = ["Service Name", "Description", "Member Of Service Groups"]
    write_header(ws, headers)

    for i, s in enumerate(rows, start=2):
        fill_row(ws, i, [
            get_val(s, "cn"),
            get_val(s, "description"),
            get_list(s, "memberof_hbacservicegroup"),
        ])

    apply_table(ws, "HBACServices", len(headers))
    auto_width(ws)


def build_hbac_service_groups(wb, data):
    rows = get_results(data, "hbac_service_groups")
    if rows is None:
        error_sheet(wb, "HBAC Svc Groups", "HBAC service group data could not be fetched from IPA.")
        return

    ws = wb.create_sheet(title="HBAC Svc Groups")
    headers = ["Group Name", "Description", "Member Services"]
    write_header(ws, headers)

    for i, g in enumerate(rows, start=2):
        fill_row(ws, i, [
            get_val(g, "cn"),
            get_val(g, "description"),
            get_list(g, "member_hbacservice"),
        ])

    apply_table(ws, "HBACSvcGroups", len(headers))
    auto_width(ws)


def build_sudo_commands(wb, data):
    rows = get_results(data, "sudo_commands")
    if rows is None:
        error_sheet(wb, "Sudo Commands", "Sudo command data could not be fetched from IPA.")
        return

    ws = wb.create_sheet(title="Sudo Commands")
    headers = ["Command", "Description", "Member Of Command Groups"]
    write_header(ws, headers)

    for i, c in enumerate(rows, start=2):
        fill_row(ws, i, [
            get_val(c, "sudocmd"),
            get_val(c, "description"),
            get_list(c, "memberof_sudocmdgroup"),
        ])

    apply_table(ws, "SudoCommands", len(headers))
    auto_width(ws)


def build_sudo_cmd_groups(wb, data):
    rows = get_results(data, "sudo_cmd_groups")
    if rows is None:
        error_sheet(wb, "Sudo Cmd Groups", "Sudo command group data could not be fetched from IPA.")
        return

    ws = wb.create_sheet(title="Sudo Cmd Groups")
    headers = ["Group Name", "Description", "Member Commands"]
    write_header(ws, headers)

    for i, g in enumerate(rows, start=2):
        fill_row(ws, i, [
            get_val(g, "cn"),
            get_val(g, "description"),
            get_list(g, "member_sudocmd"),
        ])

    apply_table(ws, "SudoCmdGroups", len(headers))
    auto_width(ws)


def build_automember(wb, data):
    group_rows = get_results(data, "automember_groups")
    hostgroup_rows = get_results(data, "automember_hostgroups")

    if group_rows is None and hostgroup_rows is None:
        error_sheet(wb, "Automember Rules", "Automember data could not be fetched from IPA.")
        return

    ws = wb.create_sheet(title="Automember Rules")
    headers = [
        "Rule Name", "Type", "Default Group",
        "Inclusive Conditions", "Exclusive Conditions",
    ]
    write_header(ws, headers)

    row_idx = 2
    for r in (group_rows or []):
        fill_row(ws, row_idx, [
            get_val(r, "cn"),
            "Group",
            get_val(r, "automemberdefaultgroup"),
            get_list(r, "automemberinclusiveregex"),
            get_list(r, "automemberexclusiveregex"),
        ])
        row_idx += 1

    for r in (hostgroup_rows or []):
        fill_row(ws, row_idx, [
            get_val(r, "cn"),
            "Host Group",
            get_val(r, "automemberdefaultgroup"),
            get_list(r, "automemberinclusiveregex"),
            get_list(r, "automemberexclusiveregex"),
        ])
        row_idx += 1

    apply_table(ws, "AutomemberRules", len(headers))
    auto_width(ws)


# ── Summary sheet ─────────────────────────────────────────────────────────────

def build_summary(wb, data, stats):
    """Insert a Summary sheet at position 0."""
    ws = wb.create_sheet(title="Summary", index=0)
    ws.tab_color = "1F4E79"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45

    meta = data.get("meta", {})
    row = 1

    def section(title):
        nonlocal row
        ws.row_dimensions[row].height = 18
        c = ws.cell(row=row, column=1, value=title)
        c.fill = FILL_SECTION
        c.font = FONT_SECTION
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

    def kv(key, value, highlight=False):
        nonlocal row
        ka = ws.cell(row=row, column=1, value=key)
        va = ws.cell(row=row, column=2, value=value)
        ka.font = FONT_BOLD
        if highlight:
            ka.fill = FILL_ALERT
            va.fill = FILL_ALERT
            ka.font = FONT_ALERT
            va.font = FONT_ALERT
        row += 1

    def blank():
        nonlocal row
        row += 1

    # ── Report metadata ──────────────────────────────────────────────────────
    section("Report Information")
    kv("Generated at",  meta.get("generated_at", "-"))
    kv("IPA Domain",    meta.get("ipa_domain", "-"))
    kv("IPA Server",    meta.get("ipa_server", "-"))
    kv("Client ID",     meta.get("client_id", "-"))
    blank()

    # ── Object counts ────────────────────────────────────────────────────────
    section("Inventory")
    count_map = [
        ("users",               "Users"),
        ("groups",              "User Groups"),
        ("hosts",               "Hosts"),
        ("host_groups",         "Host Groups"),
        ("sudo_rules",          "Sudo Rules"),
        ("hbac_rules",          "HBAC Rules"),
        ("pw_policies",         "Password Policies"),
        ("roles",               "Roles"),
        ("services",            "Services"),
        ("dns_zones",           "DNS Zones"),
        ("hbac_services",       "HBAC Services"),
        ("hbac_service_groups", "HBAC Service Groups"),
        ("sudo_commands",       "Sudo Commands"),
        ("sudo_cmd_groups",     "Sudo Command Groups"),
    ]
    for key, label in count_map:
        results = get_results(data, key)
        if results is not None:
            kv(label, len(results))
    blank()

    # ── Automember ───────────────────────────────────────────────────────────
    ag = get_results(data, "automember_groups")
    ah = get_results(data, "automember_hostgroups")
    if ag is not None or ah is not None:
        kv("Automember Rules (groups)",     len(ag) if ag is not None else "N/A")
        kv("Automember Rules (hostgroups)", len(ah) if ah is not None else "N/A")
        blank()

    # ── Alerts ───────────────────────────────────────────────────────────────
    section("Alerts")
    disabled_count, expired_count, expiring_count = stats

    has_alerts = False
    if disabled_count > 0:
        kv(f"Disabled users", disabled_count, highlight=True)
        has_alerts = True
    if expired_count > 0:
        kv(f"Users with expired password", expired_count, highlight=True)
        has_alerts = True
    if expiring_count > 0:
        kv(f"Users with password expiring in <30 days", expiring_count, highlight=False)
        has_alerts = True

    # Rules without members
    sudo_rows = get_results(data, "sudo_rules")
    if sudo_rows:
        empty_sudo = [r for r in sudo_rows if not r.get("memberuser_user") and not r.get("memberuser_group")]
        if empty_sudo:
            kv("Sudo rules without user members", len(empty_sudo))
            has_alerts = True

    hbac_rows = get_results(data, "hbac_rules")
    if hbac_rows:
        empty_hbac = [r for r in hbac_rows if not r.get("memberuser_user") and not r.get("memberuser_group")]
        if empty_hbac:
            kv("HBAC rules without user members", len(empty_hbac))
            has_alerts = True

    if not has_alerts:
        kv("Status", "No alerts detected")

    # Add a thin border under each row for readability
    for r in range(1, row):
        for col in range(1, 3):
            ws.cell(row=r, column=col).border = THIN_BORDER


# ── Sheet dispatch ────────────────────────────────────────────────────────────

SHEET_BUILDERS = {
    "users":               build_users,
    "groups":              build_groups,
    "hosts":               build_hosts,
    "host_groups":         build_host_groups,
    "sudo_rules":          build_sudo_rules,
    "hbac_rules":          build_hbac_rules,
    "pw_policies":         build_pw_policies,
    "roles":               build_roles,
    "services":            build_services,
    "dns_zones":           build_dns_zones,
    "hbac_services":       build_hbac_services,
    "hbac_service_groups": build_hbac_service_groups,
    "sudo_commands":       build_sudo_commands,
    "sudo_command_groups": build_sudo_cmd_groups,
    "automember":          build_automember,
}

ALL_SHEETS_ORDER = [
    "users", "groups", "hosts", "host_groups",
    "sudo_rules", "hbac_rules", "pw_policies",
    "roles", "services", "dns_zones",
    "hbac_services", "hbac_service_groups",
    "sudo_commands", "sudo_command_groups",
    "automember",
]


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Generate XLSX report from FreeIPA JSON data.")
    parser.add_argument("input_json",  help="Path to the combined IPA JSON data file")
    parser.add_argument("output_xlsx", help="Path to write the output XLSX file")
    args = parser.parse_args()

    # Load data
    try:
        with open(args.input_json, "r", encoding="utf-8") as f:
            data = json.load(f)
    except (json.JSONDecodeError, OSError) as exc:
        print(f"ERROR: Cannot read input JSON: {exc}", file=sys.stderr)
        sys.exit(1)

    # Determine which sheets to build (from enabled_sheets in JSON meta or all)
    enabled = data.get("meta", {}).get("enabled_sheets") or ALL_SHEETS_ORDER
    sheets_to_build = [s for s in ALL_SHEETS_ORDER if s in enabled]

    wb = Workbook()
    wb.remove(wb.active)  # Remove default empty sheet

    # Build data sheets
    user_stats = (0, 0, 0)
    for sheet_key in sheets_to_build:
        builder = SHEET_BUILDERS.get(sheet_key)
        if builder:
            result = builder(wb, data)
            if sheet_key == "users" and result:
                user_stats = result

    # Build summary sheet (inserted at index 0)
    build_summary(wb, data, user_stats)

    # Save
    try:
        wb.save(args.output_xlsx)
        meta = data.get("meta", {})
        print(f"Report saved: {args.output_xlsx} "
              f"({meta.get('ipa_domain', '?')} / {len(wb.sheetnames)} sheets)")
    except OSError as exc:
        print(f"ERROR: Cannot write output XLSX: {exc}", file=sys.stderr)
        sys.exit(1)


if __name__ == "__main__":
    main()
